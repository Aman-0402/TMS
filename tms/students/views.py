from openpyxl import load_workbook
from django.db import transaction
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import BasePermission, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from accounts.mixins import AuditLogMixin, RoleScopedQuerysetMixin, SoftDeleteMixin
from accounts.models import AuditLog, Manager
from accounts.permissions import IsAdmin, IsManager, IsTrainer

from batch.models import Batch
from labs.models import Lab
from trainers.models import Trainer
from .models import Student
from .serializers import StudentSerializer


class DenyAllPermission(BasePermission):
    def has_permission(self, request, view):
        return False


class StudentViewSet(SoftDeleteMixin, AuditLogMixin, RoleScopedQuerysetMixin, viewsets.ModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    manager_lookup = "batch__manager__user"
    trainer_lookup = "lab__trainer__user"

    def get_permissions(self):
        self.debug_request_user("get_permissions")
        role_permission_map = {
            "ADMIN": IsAdmin,
            "MANAGER": IsManager,
            "TRAINER": IsTrainer,
        }
        role = getattr(self.request.user, "role", None)
        permission_class = role_permission_map.get(role, DenyAllPermission)

        if self.action in {
            "list",
            "retrieve",
            "create",
            "update",
            "partial_update",
            "destroy",
            "restore",
            "upload_excel",
        }:
            return [IsAuthenticated(), permission_class()]

        return [IsAuthenticated()]

    def get_base_queryset(self):
        queryset = (
            Student.objects.select_related("batch", "lab", "lab__trainer", "lab__trainer__user")
            .prefetch_related("results")
            .order_by("batch__name", "name")
        )

        batch_id = self.request.query_params.get("batch")
        trainer_id = self.request.query_params.get("trainer")
        lab_id = self.request.query_params.get("lab")

        if batch_id:
            queryset = queryset.filter(batch_id=batch_id)
        if trainer_id:
            queryset = queryset.filter(lab__trainer_id=trainer_id)
        if lab_id:
            queryset = queryset.filter(lab_id=lab_id)

        return queryset

    def perform_create(self, serializer):
        self._validate_write_scope(serializer)
        super().perform_create(serializer)

    def perform_update(self, serializer):
        self._assert_user_can_modify_student(serializer.instance)
        self._validate_write_scope(serializer)
        super().perform_update(serializer)

    def perform_destroy(self, instance):
        self._assert_user_can_modify_student(instance)
        super().perform_destroy(instance)

    def _assert_user_can_modify_student(self, student):
        user = self.request.user

        if user.role == "ADMIN":
            return

        if user.role == "MANAGER":
            if not Manager.objects.filter(user=user, batch=student.batch).exists():
                raise PermissionDenied("Managers can only modify students in their assigned batch.")
            return

        if user.role == "TRAINER":
            if not student.lab or student.lab.trainer.user_id != user.id:
                raise PermissionDenied("Trainers can only modify students in their assigned lab.")
            return

        raise PermissionDenied("You do not have permission to modify this student.")

    def _validate_write_scope(self, serializer):
        user = self.request.user
        batch = serializer.validated_data.get("batch", getattr(serializer.instance, "batch", None))
        lab = serializer.validated_data.get("lab", getattr(serializer.instance, "lab", None))

        if user.role == "ADMIN":
            return

        if user.role == "MANAGER":
            if not batch or not Manager.objects.filter(user=user, batch=batch).exists():
                raise PermissionDenied("Managers can only manage students in their assigned batch.")
            return

        if user.role == "TRAINER":
            if not lab or lab.trainer.user_id != user.id:
                raise PermissionDenied("Trainers can only manage students in their assigned lab.")
            return

        raise PermissionDenied("You do not have permission to manage students.")

    @action(detail=False, methods=["post"], url_path="upload-excel")
    def upload_excel(self, request):
        uploaded_file = request.FILES.get("file")
        batch_id = request.data.get("batch")

        if not uploaded_file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        if not batch_id:
            return Response({"error": "Batch is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not uploaded_file.name.lower().endswith((".xlsx", ".xls")):
            return Response(
                {"error": "Wrong file format. Please upload an Excel file (.xlsx or .xls)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        batch = Batch.objects.filter(pk=batch_id, is_deleted=False).first()
        if not batch:
            return Response({"error": "Batch not found"}, status=status.HTTP_404_NOT_FOUND)

        self._validate_batch_upload_scope(batch)

        try:
            workbook = load_workbook(uploaded_file, data_only=True)
            worksheet = workbook.active
        except Exception:
            return Response(
                {"error": "Unable to read the uploaded Excel file. Please check the format and try again."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Extract rows as list of dicts
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows or len(rows) == 0:
            return Response(
                {"error": "The uploaded Excel file is empty."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Parse header row (first row)
        header_raw = rows[0]
        if not header_raw or all(cell is None for cell in header_raw):
            return Response(
                {"error": "The uploaded Excel file has no headers."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Normalize column names
        headers = [str(cell).strip() if cell is not None else "" for cell in header_raw]
        headers = [h for h in headers if h]  # Remove empty headers

        # Check required columns
        required_columns = {"UG Number", "Name", "Department", "Lab"}
        header_set = set(headers)
        missing_columns = required_columns - header_set
        if missing_columns:
            return Response(
                {
                    "error": "Missing required columns.",
                    "missing_columns": sorted(missing_columns),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Convert rows to list of dicts
        data_rows = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue  # Skip empty rows
            row_dict = {}
            for i, header in enumerate(headers):
                row_dict[header] = row[i] if i < len(row) else None
            data_rows.append(row_dict)

        if not data_rows:
            return Response(
                {"error": "The uploaded Excel file has no data rows."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Extract and normalize UG numbers
        normalized_ug_numbers = [
            self._normalize_excel_value(row.get("UG Number"))
            for row in data_rows
            if self._normalize_excel_value(row.get("UG Number"))
        ]

        # Check for duplicates in file
        duplicate_ug_numbers_in_file = sorted(
            {
                ug_number
                for ug_number in normalized_ug_numbers
                if normalized_ug_numbers.count(ug_number) > 1
            }
        )
        if duplicate_ug_numbers_in_file:
            return Response(
                {
                    "error": "Duplicate UG numbers found in the uploaded file.",
                    "duplicate_ug_numbers": duplicate_ug_numbers_in_file,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        labs_by_name = {
            str(lab.name).strip(): lab
            for lab in Lab.objects.select_related("trainer", "trainer__user").filter(batch=batch)
        }

        # Map existing students by UG number, including their current course
        existing_students = {
            s.ug_number: s
            for s in Student.objects.select_related("batch__course").filter(
                ug_number__in=normalized_ug_numbers
            )
        }

        students_to_create = []
        students_to_reenroll = []
        created_ug_numbers = []
        same_course_duplicates = []
        processed_in_file = set()

        with transaction.atomic():
            for row_number, row_data in enumerate(data_rows, start=2):
                ug_number = self._normalize_excel_value(row_data.get("UG Number"))
                name = self._normalize_excel_value(row_data.get("Name"))
                department = self._normalize_excel_value(row_data.get("Department"))
                lab_name = self._normalize_excel_value(row_data.get("Lab"))
                email = self._normalize_excel_value(row_data.get("Email"))
                phone = self._normalize_excel_value(row_data.get("Phone"))

                if not ug_number or not name or not department or not lab_name:
                    return Response(
                        {"error": f"Required value missing in row {row_number}."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                if ug_number in processed_in_file:
                    continue
                processed_in_file.add(ug_number)

                lab = labs_by_name.get(lab_name)
                if not lab:
                    lab = self._get_or_create_upload_lab(batch, lab_name, labs_by_name)

                if not lab:
                    return Response(
                        {"error": f"Unable to assign or create lab '{lab_name}' for batch '{batch.name}'."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                self._validate_lab_upload_scope(lab)

                existing = existing_students.get(ug_number)
                if existing:
                    # Same course → duplicate, skip
                    if existing.batch.course_id == batch.course_id:
                        same_course_duplicates.append({
                            "ug_number": ug_number,
                            "name": existing.name,
                            "current_batch": existing.batch.name,
                        })
                    else:
                        # Different course → re-enroll (move to new batch/lab)
                        students_to_reenroll.append((existing, batch, lab, name, department, email, phone))
                    continue

                students_to_create.append(
                    Student(
                        ug_number=ug_number,
                        name=name,
                        department=department,
                        email=email,
                        phone=phone,
                        batch=batch,
                        lab=lab,
                    )
                )
                created_ug_numbers.append(ug_number)

            # Bulk-create new students
            Student.objects.bulk_create(students_to_create, batch_size=1000)

            # Re-enroll existing students
            reenrolled_ids = []
            for student, new_batch, new_lab, new_name, new_dept, new_email, new_phone in students_to_reenroll:
                student.batch = new_batch
                student.lab   = new_lab
                student.name  = new_name or student.name
                student.department = new_dept or student.department
                student.email = new_email or student.email
                student.phone = new_phone or student.phone
                student.is_deleted = False
                student.deleted_at = None
                student.save()
                reenrolled_ids.append(student.pk)

            # Audit log for new creates
            created_students = list(
                Student.objects.filter(ug_number__in=created_ug_numbers).select_related("batch")
            )
            audit_logs = [
                AuditLog(
                    user=request.user,
                    action=AuditLog.Action.CREATE,
                    model_name="Student",
                    object_id=student.pk,
                    description=f"Created Student via Excel upload: {student}",
                )
                for student in created_students
            ]
            # Audit log for re-enrollments
            for pk in reenrolled_ids:
                audit_logs.append(
                    AuditLog(
                        user=request.user,
                        action=AuditLog.Action.UPDATE,
                        model_name="Student",
                        object_id=pk,
                        description=f"Re-enrolled Student (different course) via Excel upload",
                    )
                )
            AuditLog.objects.bulk_create(audit_logs, batch_size=1000)

        return Response(
            {
                "message": (
                    f"{len(students_to_create)} student(s) created, "
                    f"{len(students_to_reenroll)} re-enrolled, "
                    f"{len(same_course_duplicates)} same-course duplicate(s) skipped."
                ),
                "created_count": len(students_to_create),
                "reenrolled_count": len(students_to_reenroll),
                "same_course_duplicates": same_course_duplicates,
            },
            status=status.HTTP_201_CREATED,
        )

    def _validate_batch_upload_scope(self, batch):
        user = self.request.user

        if user.role == "ADMIN":
            return

        if user.role == "MANAGER":
            if not Manager.objects.filter(user=user, batch=batch).exists():
                raise PermissionDenied("Managers can only upload students to their assigned batch.")
            return

        if user.role == "TRAINER":
            if not Lab.objects.filter(batch=batch, trainer__user=user).exists():
                raise PermissionDenied("Trainers can only upload students to batches with their labs.")
            return

        raise PermissionDenied("You do not have permission to upload students.")

    def _validate_lab_upload_scope(self, lab):
        user = self.request.user

        if user.role == "TRAINER" and lab.trainer.user_id != user.id:
            raise PermissionDenied("Trainers can only assign students to their own labs.")

    def _get_or_create_upload_lab(self, batch, lab_name, labs_by_name):
        trainer = self._resolve_upload_lab_trainer(batch)
        if self.request.user.role == "TRAINER" and not trainer:
            return None

        lab, _ = Lab.objects.get_or_create(
            name=lab_name,
            batch=batch,
            defaults={"trainer": trainer},
        )
        labs_by_name[lab_name] = lab
        return lab

    def _resolve_upload_lab_trainer(self, batch):
        user = self.request.user

        if user.role == "TRAINER":
            trainer = getattr(user, "trainer_profile", None)
            if trainer and trainer.batch_id == batch.id:
                return trainer
            return None

        return Trainer.objects.filter(batch=batch).order_by("id").first()

    def _normalize_excel_value(self, value):
        """
        Normalize Excel cell values to clean strings.
        Handles None, NaN, floats, and whitespace.
        """
        if value is None:
            return ""
        # Check if it's a float that should be converted to int
        if isinstance(value, float):
            if value != value:  # NaN check (NaN != NaN)
                return ""
            if value.is_integer():
                return str(int(value))
        return str(value).strip()
